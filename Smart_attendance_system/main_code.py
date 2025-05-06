import cv2
import numpy as np
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Initialize OpenCV's face detector (Haar Cascade)
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Initialize OpenCV's LBPH face recognizer
recognizer = cv2.face.LBPHFaceRecognizer_create()

# Folder containing known people's images (training data)
image_folder = 'images'
attendance_file = 'attendance.xlsx'

# Track people who have already been marked present today
attendance_today = set()

# Confidence threshold for unknown detection
CONFIDENCE_THRESHOLD = 70  # Below this value is considered a good match

# Email configuration
EMAIL_ADDRESS = "hy66331@gmail.com"
EMAIL_PASSWORD = "reqq bsve rvbx wztn"
TO_EMAIL = "siddhipawar2102@gmail.com"

# Function to prepare the training data (images)
def prepare_training_data(image_folder):
    faces = []
    labels = []
    names = {}

    for label, filename in enumerate(os.listdir(image_folder)):
        if filename.endswith('.jpg') or filename.endswith('.png'):
            image_path = os.path.join(image_folder, filename)
            image = cv2.imread(image_path)
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

            faces_in_image = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

            for (x, y, w, h) in faces_in_image:
                faces.append(gray[y:y + h, x:x + w])
                labels.append(label)
                names[label] = filename.split('.')[0]

    return faces, labels, names

# Load and train the face recognizer
faces, labels, names = prepare_training_data(image_folder)
recognizer.train(faces, np.array(labels))

# Function to mark attendance in Excel file
def mark_attendance(lecture_name, name):
    if name in attendance_today:
        return False

    if os.path.exists(attendance_file):
        wb = load_workbook(attendance_file)
    else:
        wb = Workbook()

    if lecture_name not in wb.sheetnames:
        sheet = wb.create_sheet(lecture_name)
        sheet.append(['Name', 'Date & Time'])
    else:
        sheet = wb[lecture_name]

    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet.append([name, current_time])
    wb.save(attendance_file)

    attendance_today.add(name)
    return True

# Function to send an email notification
def send_email(name, lecture_name):
    subject = f"Attendance Marked for {lecture_name}"
    body = f"Attendance has been successfully marked for {name} in lecture {lecture_name} on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    msg = MIMEMultipart()
    msg['From'] = EMAIL_ADDRESS
    msg['To'] = TO_EMAIL
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.send_message(msg)
        print(f"Email notification sent for {name}.")
    except Exception as e:
        print(f"Failed to send email: {e}")

# Prompt for lecture name
lecture_name = input("Enter the lecture name: ")

# Start the webcam stream
cap = cv2.VideoCapture(0)
last_recognized_time = {}

while True:
    ret, frame = cap.read()
    if not ret:
        break

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    faces_in_frame = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

    for (x, y, w, h) in faces_in_frame:
        face_region = gray[y:y + h, x:x + w]
        label, confidence = recognizer.predict(face_region)

        name = names.get(label, "Unknown")

        if confidence > CONFIDENCE_THRESHOLD:
            name = "Unknown"

        current_time = time.time()

        if name != "Unknown" and (name not in last_recognized_time or current_time - last_recognized_time[name] > 1):
            if mark_attendance(lecture_name, name):
                print(f"Attendance marked for {name}.")
                send_email(name, lecture_name)
            else:
                print(f"{name} is already marked present.")
            last_recognized_time[name] = current_time
        elif name == "Unknown":
            print("Unknown person detected.")

        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
        cv2.putText(frame, name, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

    cv2.imshow('Face Recognition Attendance System', frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()
