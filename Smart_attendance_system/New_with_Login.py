from flask import Flask, render_template, Response, request, jsonify, send_from_directory, redirect, url_for, session
import cv2
import numpy as np
import os
import base64
from datetime import datetime
from openpyxl import Workbook, load_workbook
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import threading
import shutil
import secrets

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)  # Generate a random secret key for session

# Global variables
camera = None
frame_counter = 0
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
recognizer = cv2.face.LBPHFaceRecognizer_create()
attendance_today = set()
current_lecture = None
processing = False
last_recognized_time = {}
latest_attendance = {"name": None, "time": None, "image": None}

# Email configuration
EMAIL_ADDRESS = "hy66331@gmail.com"
EMAIL_PASSWORD = "xxxxxxxxxxxxxx"
TO_EMAIL = "siddhipawar2102@gmail.com"

# Admin credentials (for simplicity, this is hardcoded; in production, use hashed passwords)
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"

# Confidence threshold for unknown detection
CONFIDENCE_THRESHOLD = 70  # Below this value is considered a good match

# File paths
image_folder = 'images'
static_images = 'static/student_images'
attendance_file = 'attendance.xlsx'
lectures_file = 'lectures.txt'

# Ensure directories exist
for directory in [image_folder, static_images]:
    if not os.path.exists(directory):
        os.makedirs(directory)

# Function to get all lectures
def get_all_lectures():
    if not os.path.exists(lectures_file):
        with open(lectures_file, 'w') as f:
            f.write("Database Management\nArtificial Intelligence\nMachine Learning\nWeb Development\nCyber Security")
    
    with open(lectures_file, 'r') as f:
        lectures = [line.strip() for line in f.readlines() if line.strip()]
    
    return lectures

# Function to add a new lecture
def add_lecture(lecture_name):
    lectures = get_all_lectures()
    if lecture_name not in lectures:
        with open(lectures_file, 'a') as f:
            f.write(f"\n{lecture_name}")
        return True
    return False

# Function to prepare the training data (images)
def prepare_training_data(image_folder):
    faces = []
    labels = []
    names = {}

    if not os.path.exists(image_folder):
        os.makedirs(image_folder)
        print(f"Created directory: {image_folder}")
        return faces, labels, names

    for label, filename in enumerate(os.listdir(image_folder)):
        if filename.endswith('.jpg') or filename.endswith('.png'):
            image_path = os.path.join(image_folder, filename)
            image = cv2.imread(image_path)
            if image is None:
                print(f"Failed to load image: {image_path}")
                continue

            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            faces_in_image = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

            for (x, y, w, h) in faces_in_image:
                faces.append(gray[y:y + h, x:x + w])
                labels.append(label)
                names[label] = filename.split('.')[0]

    return faces, labels, names


# Initialize face recognition on startup
def initialize_face_recognition():
    faces, labels, names = prepare_training_data(image_folder)
    if faces and labels:
        recognizer.train(faces, np.array(labels))
    app.config['names'] = names


# Function to mark attendance in Excel file
def mark_attendance(lecture_name, name):
    global latest_attendance

    if name in attendance_today:
        return False

    if os.path.exists(attendance_file):
        wb = load_workbook(attendance_file)
    else:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    if lecture_name not in wb.sheetnames:
        sheet = wb.create_sheet(lecture_name)
        sheet.append(['Name', 'Date & Time'])
    else:
        sheet = wb[lecture_name]

    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet.append([name, current_time])
    wb.save(attendance_file)

    attendance_today.add(name)

    # Update latest attendance information and copy student image to static folder
    source_image = os.path.join(image_folder, f"{name}.jpg")
    if os.path.exists(source_image):
        target_image = os.path.join(static_images, f"{name}.jpg")
        shutil.copy(source_image, target_image)
        latest_attendance = {
            "name": name,
            "time": current_time,
            "image": f"student_images/{name}.jpg"
        }

    return True


# Function to send an email notification
def send_email(name, lecture_name):
    subject = f"Attendance Marked for {lecture_name}"
    body = f"Attendance has been successfully marked for {name} in lecture {lecture_name} on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    msg = MIMEMultipart()
    msg['From'] = EMAIL_ADDRESS
    msg['To'] = TO_EMAIL
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.send_message(msg)
        print(f"Email notification sent for {name}.")
    except Exception as e:
        print(f"Failed to send email: {e}")


def generate_frames():
    global camera, processing, frame_counter, current_lecture

    if camera is None:
        camera = cv2.VideoCapture(0)

    while True:
        success, frame = camera.read()
        if not success:
            break
        else:
            # Only process every 10th frame to reduce CPU usage
            frame_counter += 1
            if frame_counter % 10 == 0 and current_lecture and not processing:
                threading.Thread(target=process_frame, args=(frame.copy(), current_lecture)).start()

            # Draw UI elements on the frame
            cv2.putText(frame, f"Lecture: {current_lecture if current_lecture else 'Not Set'}",
                        (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

            ret, buffer = cv2.imencode('.jpg', frame)
            frame = buffer.tobytes()

            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')


def process_frame(frame, lecture_name):
    global processing, last_recognized_time

    processing = True
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces_in_frame = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

    for (x, y, w, h) in faces_in_frame:
        face_region = gray[y:y + h, x:x + w]

        try:
            label, confidence = recognizer.predict(face_region)
            name = app.config['names'].get(label, "Unknown")

            if confidence > CONFIDENCE_THRESHOLD:
                name = "Unknown"

            current_time = time.time()
            if name != "Unknown" and (
                    name not in last_recognized_time or current_time - last_recognized_time[name] > 30):
                if mark_attendance(lecture_name, name):
                    print(f"Attendance marked for {name}.")
                    send_email(name, lecture_name)
                else:
                    print(f"{name} is already marked present.")
                last_recognized_time[name] = current_time

            # Draw rectangle and name on the frame
            cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
            cv2.putText(frame, f"{name} ({confidence:.1f})",
                        (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

        except Exception as e:
            print(f"Error during face recognition: {e}")

    processing = False

# Login route
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        if 'logged_in' in session:
            return redirect(url_for('index'))
        return render_template('login.html')
    
    if request.method == 'POST':
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['logged_in'] = True
            return jsonify({"status": "success"})
        else:
            return jsonify({"status": "error", "message": "Invalid username or password"})

# Logout route
@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

# Route decorator to check if user is logged in
def login_required(f):
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@app.route('/')
@login_required
def index():
    # Initialize face recognition on first request
    if 'names' not in app.config:
        initialize_face_recognition()
    lectures = get_all_lectures()
    return render_template('index.html', lectures=lectures)


@app.route('/add_student')
@login_required
def add_student():
    if 'names' not in app.config:
        initialize_face_recognition()
    return render_template('add_student.html')


@app.route('/attendance_records')
@login_required
def attendance_records():
    if 'names' not in app.config:
        initialize_face_recognition()
    return render_template('attendance_records.html')


@app.route('/video_feed')
@login_required
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')


@app.route('/set_lecture', methods=['POST'])
@login_required
def set_lecture():
    global current_lecture, attendance_today

    data = request.get_json()
    current_lecture = data.get('lecture_name')
    attendance_today = set()  # Reset attendance for new lecture

    return jsonify({"status": "success", "message": f"Lecture set to {current_lecture}"})


@app.route('/get_latest_attendance', methods=['GET'])
@login_required
def get_latest_attendance():
    return jsonify({
        "status": "success",
        "name": latest_attendance.get("name"),
        "time": latest_attendance.get("time"),
        "image": latest_attendance.get("image")
    })


@app.route('/upload_image', methods=['POST'])
@login_required
def upload_image():
    if 'image' not in request.files or 'name' not in request.form:
        return jsonify({"status": "error", "message": "Missing image or name"})

    image_file = request.files['image']
    name = request.form['name']

    if image_file.filename == '':
        return jsonify({"status": "error", "message": "No selected image"})

    if not os.path.exists(image_folder):
        os.makedirs(image_folder)

    image_path = os.path.join(image_folder, f"{name}.jpg")
    image_file.save(image_path)

    # Also save to static folder for display
    static_image_path = os.path.join(static_images, f"{name}.jpg")
    image_file.seek(0)  # Reset file pointer to beginning
    image_file.save(static_image_path)

    # Retrain the recognizer with the new image
    faces, labels, names = prepare_training_data(image_folder)
    if faces and labels:
        recognizer.train(faces, np.array(labels))
        app.config['names'] = names

    return jsonify({"status": "success", "message": f"Image for {name} uploaded successfully"})


@app.route('/get_attendance', methods=['GET'])
@login_required
def get_attendance():
    if not os.path.exists(attendance_file):
        return jsonify({"status": "error", "message": "No attendance records found"})

    wb = load_workbook(attendance_file)
    attendance_data = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        attendance_data[sheet_name] = []

        for row in list(sheet.rows)[1:]:  # Skip header row
            if len(row) >= 2:
                attendance_data[sheet_name].append({
                    "name": row[0].value,
                    "timestamp": str(row[1].value)
                })

    return jsonify({"status": "success", "data": attendance_data})


@app.route('/get_student_list', methods=['GET'])
@login_required
def get_student_list():
    students = []

    if os.path.exists(image_folder):
        for filename in os.listdir(image_folder):
            if filename.endswith('.jpg') or filename.endswith('.png'):
                name = filename.split('.')[0]
                students.append({
                    "name": name,
                    "image": f"student_images/{filename}"
                })

    return jsonify({"status": "success", "data": students})


@app.route('/get_lectures', methods=['GET'])
@login_required
def get_lectures():
    lectures = get_all_lectures()
    return jsonify({"status": "success", "data": lectures})


@app.route('/add_new_lecture', methods=['POST'])
@login_required
def add_new_lecture():
    data = request.get_json()
    lecture_name = data.get('lecture_name')
    
    if not lecture_name:
        return jsonify({"status": "error", "message": "Lecture name is required"})
    
    if add_lecture(lecture_name):
        return jsonify({"status": "success", "message": f"Lecture '{lecture_name}' added successfully"})
    else:
        return jsonify({"status": "error", "message": f"Lecture '{lecture_name}' already exists"})


@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory('static', filename)

# new from here
#
# @app.route('/capture_for_add_student', methods=['POST'])
# @login_required
# def capture_for_add_student():
#     global camera
#
#     if camera is None:
#         camera = cv2.VideoCapture(0)
#
#     success, frame = camera.read()
#     if not success:
#         return jsonify({"status": "error", "message": "Failed to capture image"})
#
#     # Convert the captured frame to JPEG
#     ret, buffer = cv2.imencode('.jpg', frame)
#     if not ret:
#         return jsonify({"status": "error", "message": "Failed to encode image"})
#
#     # Convert to base64 for sending to frontend
#     image_data = "data:image/jpeg;base64," + base64.b64encode(buffer).decode('utf-8')
#
#     return jsonify({
#         "status": "success",
#         "message": "Image captured",
#         "image_data": image_data
#     })
# to here

if __name__ == '__main__':
    # Initialize the app with the recognition model
    initialize_face_recognition()
    app.run(debug=True)
